#!/usr/bin/env python3
"""
Import des membres depuis le fichier Excel vers Supabase.
Usage: python3 scripts/import_membres.py

Nécessite: pip install openpyxl supabase
Variables d'environnement requises:
  SUPABASE_URL
  SUPABASE_SERVICE_KEY  (service role, pas anon key)
"""

import os
import sys
import json
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("Installer openpyxl: pip install openpyxl")
    sys.exit(1)

HAS_SUPABASE = False
try:
    from supabase import create_client
    HAS_SUPABASE = True
except ImportError:
    pass

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "docs", "fidelite",
    "2024-2026 - Liste des membres.xlsx"
)

# Headers are on row 4, data starts row 5
HEADER_ROW = 4
DATA_START_ROW = 5

# AGA/AGE column mapping (column letters → event names)
AGA_COLUMNS = {
    "R": "AGA 2023 (Sept 2023)",
    "S": "AGE (9 janvier 2024)",
    "T": "AGE (1 février 2024)",
    "U": "AGE (26 février 2024)",
    "V": "AGA 2024 (oct. 2024)",
}


def col_letter_to_index(letter):
    """Convert column letter to 0-based index."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1


def parse_date(value):
    """Parse date from Excel cell value."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    # Try parsing string dates
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def parse_presence(value):
    """Parse presence value to boolean or None."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("oui", "o", "1", "x", "true", "présent", "present", "p"):
        return True
    if s in ("non", "n", "0", "false", "absent", "a", ""):
        return False
    return None


def load_excel():
    """Load members from the first sheet of the Excel file."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]  # "Liste membres 2024-25"

    members = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]

        # Skip empty rows (check nom)
        nom = row[1]  # Column B
        if not nom or (isinstance(nom, str) and not nom.strip()):
            continue

        # Parse AGA/AGE presences
        presences = {}
        for col_letter, event_name in AGA_COLUMNS.items():
            col_idx = col_letter_to_index(col_letter)
            if col_idx < len(row):
                val = parse_presence(row[col_idx])
                if val is not None:
                    presences[event_name] = val

        member = {
            "titre": str(row[0]).strip() if row[0] else None,
            "nom": str(row[1]).strip(),
            "prenom": str(row[2]).strip() if row[2] else "",
            "telephone": str(row[3]).strip() if row[3] else None,
            "email": str(row[4]).strip().lower() if row[4] else None,
            "adresse": str(row[5]).strip() if row[5] else None,
            "num_app": str(row[6]).strip() if row[6] else None,
            "ville": str(row[7]).strip() if row[7] else None,
            "code_postal": str(row[8]).strip() if row[8] else None,
            "date_adhesion": parse_date(row[9]),
            "date_renouvellement": parse_date(row[10]),
            "date_fin": parse_date(row[11]),
            "langue": str(row[13]).strip() if row[13] else None,
            "etat": str(row[14]).strip().upper() if row[14] else None,
            "type_membre": str(row[15]).strip().lower() if row[15] else None,
            "detail_type": str(row[16]).strip() if row[16] else None,
            "presences_aga": presences if presences else {},
        }

        members.append(member)

    print(f"Loaded {len(members)} members from Excel")
    return members


def upload_to_supabase(members):
    """Upload members to Supabase in batches."""
    if not HAS_SUPABASE:
        print("ERROR: pip install supabase required for upload")
        sys.exit(1)
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY")

    if not url or not key:
        print("ERROR: Set SUPABASE_URL and SUPABASE_SERVICE_KEY environment variables")
        print("  export SUPABASE_URL='https://xxx.supabase.co'")
        print("  export SUPABASE_SERVICE_KEY='eyJ...'")
        sys.exit(1)

    client = create_client(url, key)

    BATCH_SIZE = 100
    total = 0
    for i in range(0, len(members), BATCH_SIZE):
        batch = members[i:i + BATCH_SIZE]
        # Convert presences_aga dict to JSON string for Supabase
        for m in batch:
            m["presences_aga"] = json.dumps(m["presences_aga"])

        result = client.table("membres").insert(batch).execute()
        total += len(batch)
        print(f"  Inserted {total}/{len(members)}")

    print(f"Done! {total} members imported.")


def dry_run(members):
    """Print summary without uploading."""
    print(f"\n=== DRY RUN — {len(members)} membres ===")
    print(f"Sample (first 3):")
    for m in members[:3]:
        print(f"  {m['prenom']} {m['nom']} | {m['email']} | adhésion: {m['date_adhesion']} | fin: {m['date_fin']} | type: {m['type_membre']} | état: {m['etat']}")

    # Stats
    types = {}
    etats = {}
    with_email = 0
    for m in members:
        t = m["type_membre"] or "non spécifié"
        types[t] = types.get(t, 0) + 1
        e = m["etat"] or "?"
        etats[e] = etats.get(e, 0) + 1
        if m["email"]:
            with_email += 1

    print(f"\nTypes: {types}")
    print(f"États: {etats}")
    print(f"Avec email: {with_email}/{len(members)}")


if __name__ == "__main__":
    members = load_excel()

    if "--upload" in sys.argv:
        upload_to_supabase(members)
    else:
        dry_run(members)
        print("\nPour importer dans Supabase, relancer avec --upload")
        print("  python3 scripts/import_membres.py --upload")
